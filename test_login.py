import pytest
from selenium import webdriver
import openpyxl
from datetime import datetime

from login_page import LoginPage  # Import the LoginPage class from the above code

def read_test_data(file_path):
    workbook = openpyxl.load_workbook(file_path)
    sheet = workbook['LoginTest']
    data = []
    for row in sheet.iter_rows(min_row=2, values_only=True):
        data.append(row)
    return data, workbook

def write_test_result(workbook, row, result):
    sheet = workbook['LoginTest']
    sheet.cell(row=row+1, column=7, value=result)
    workbook.save('login_tests.xlsx')

@pytest.fixture
def driver():
    driver = webdriver.Chrome()
    yield driver
    driver.quit()

@pytest.mark.parametrize("test_id, username, password, date, time, tester, result", read_test_data('login_tests.xlsx')[0])
def test_login(driver, test_id, username, password, date, time, tester, result):
    driver.get("https://opensource-demo.orangehrmlive.com/web/index.php/auth/login")
    login_page = LoginPage(driver)
    
    login_page.enter_username(username)
    login_page.enter_password(password)
    login_page.click_login()

    if login_page.is_login_successful():
        test_result = 'Passed'
    else:
        test_result = 'Failed'

    workbook = read_test_data('login_tests.xlsx')[1]
    write_test_result(workbook, int(test_id), test_result)

if __name__ == "__main__":
    pytest.main()
